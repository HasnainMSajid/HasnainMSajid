import openpyxl
from openpyxl.styles import Font
#from openpyxl.reader.excel import load_workbook 
#from openpyxl import load_workbook
#from openpyxl import Workbook
#print(openpyxl.__version__)

#Took help from these videos:
    #https://www.youtube.com/watch?v=8z61LhMsyDM
    #https://www.youtube.com/watch?v=JjeSXWIX9SY


'''make a sheet'''
#wb = openpyxl.Workbook()

'''Load filename.xlsx'''
wb = openpyxl.load_workbook('energy_data.xlsx')

'''change the sheet name'''
ws = wb['Sheet1']
ws.title = 'Energies'
'''del col_3 with ----'''
ws.delete_cols(3)

'''add title in row one''''''First add a row on top'''
ws.insert_rows(1, 1)

'''adding title using enumerate, start=1 means that adding title should start with column 1'''
for col_idx, title in enumerate(('Names', 'Energies (au)', 'Erel (kcal/mol)', 'Erel_round'), start=1):
    ws.cell(row=1, column=col_idx, value=title)

'''To apply font on column one'''
for i in range(1, ws.max_column+1):
    ws.cell(row=1, column=i).font = Font(bold=True, name='Arial', size=10)


'''To apply font on column one'''
for i in range(1, ws.max_column+1):
    ws.cell(row=1, column=i).font = Font(bold=True, name='Arial', size=10)

'''Insert colum for adding formula'''
#ws.insert_cols(3)
#select rows from 2 to infinity. Here max_row+1 means till last row

for i in range(2, ws.max_row+1):
    #print(i)
    #select values of column 2 (energies)
    Energies = float(ws.cell(row=i, column=2).value)
    #print(Energies)
    #select the energy of first/AA structure
    initial = float(ws.cell(row=2, column=2).value)
    #write a formula
    Erel = (initial - Energies)*627.51
    #append formula to column 3, title Erel 
    ws.cell(row=i, column=3).value = Erel  

'''Add round function'''
for i in range(2, ws.max_row+1):
    Erel = ws.cell(row=i, column=3).value
    formula = round(Erel, 0)
    ws.cell(row=i, column=4).value = formula
'''save a sheet'''
wb.save('Excel_sheet_Final.xlsx')